# streamlit_app.py
# ---------------------------------------------------------
# Taiga vs TRAD TCO – per-unit scaling, TRAD €/m² investment,
# TRAD costs scaled by number of rooms (qty) and years.
# Includes Leasing calculator that uses the same WACC definition.
# ---------------------------------------------------------

import streamlit as st, base64
import pandas as pd
import io
import matplotlib.pyplot as plt
import leasing_calc as lc
from importlib import reload
from pathlib import Path

from tco_v2 import TCOInputs, CyclePoint, compute_tco, yearly_breakdown
import proposal_doc  # keep as module, so reload works

st.set_page_config(page_title="Taiga Calculator", layout="wide")

css = Path("style.css").read_text()

# -------------------- Helpers --------------------

def init_state():
    ss = st.session_state

    # Shared
    ss.setdefault("years", 5)
    ss.setdefault("wacc", 0.05)          # annual WACC (decimal), used everywhere incl. Leasing
    ss.setdefault("area_m2", 4.0)
    ss.setdefault("kwh_m2yr", 105.0)
    ss.setdefault("elec_price", 0.15)
    ss.setdefault("cycle_year", 5)       # Taiga only
    ss.setdefault("area_from_products", None)  # computed from selector

    # Taiga cycle table (only for Taiga)
    if "cycle_df" not in ss:
        ss.cycle_df = pd.DataFrame({
            "year": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            "value_pct": [50, 50, 50, 40, 35, 30, 25, 20, 15, 10],
        })

    # ---------- Taiga per-unit defaults ----------
    ss.setdefault("taiga_list_price", 10_000.0)
    ss.setdefault("taiga_occ_rate", 0.35)
    ss.setdefault("taiga_standby", 0.05)

    ss.setdefault("taiga_commissioning_cost_unit", 950.0)   # €/unit
    ss.setdefault("taiga_maint_total_unit", 150.0)          # €/unit annually
    ss.setdefault("taiga_dt_rate", 15.0)                    # €/h
    ss.setdefault("taiga_dt_install_h_unit", 8.0)           # h/unit
    ss.setdefault("taiga_dt_maint_h_total_unit", 1.5)       # h/unit annually

    ss.setdefault("taiga_commissioning_year", 1)
    ss.setdefault("taiga_eol_cost", 0.0)                    # total (not scaled)

    ss.setdefault("taiga_total_qty", 0)                     # total selected qty

    # ---------- TRAD defaults ----------
    ss.setdefault("trad_list_price", 10_000.0)              # not used in calc
    ss.setdefault("trad_price_per_m2", 2500.0)              # investment €/m²

    ss.setdefault("trad_commissioning_cost_unit", 1500.0)   # €/room/year
    ss.setdefault("trad_commissioning_year", 1)
    ss.setdefault("trad_maint_total_unit", 500.0)           # €/room annually

    ss.setdefault("trad_dt_rate", 15.0)                     # €/h
    ss.setdefault("trad_dt_install_h_unit", 80.0)           # h/room
    ss.setdefault("trad_dt_maint_h_total_unit", 5.0)        # h/room annually

    ss.setdefault("trad_eol_pct", 0.20)                     # 20% of investment
    ss.setdefault("trad_run_frac", 0.90)

    # Taiga product selector state
    ss.setdefault("override_price", False)
    ss.setdefault("override_area", False)
    if "price_df" not in ss:
        ss.price_df = pd.DataFrame({
            "code": ["LB1", "LB2", "LB3", "LB5", "LB7", "PIC", "FL10", "FL12", "FL14", "FL21", "FL25", "FL28"],
            "name": ["Taiga LB1", "Taiga LB2", "Taiga LB3", "Taiga LB5", "Taiga LB7", "Picea", "Flex10", "Flex12", "Flex14", "Flex21", "Flex25", "Flex28"],
            "unit_price_eur": [9900, 14900, 15900, 21900, 25900, 18900, 44540, 50290, 57520, 67060, 74080, 82220],
            "area_m2": [1, 2, 3, 5, 7, 3, 10, 12, 14, 21, 25, 28],
            "qty": [0]*12,
        })

def taiga_price_list_ui():
    """Upload/select Taiga products, edit qty; compute total price, total area, total qty."""
    st.markdown("**Taiga product selector (optional)**")
    st.caption("Upload a price list or use the default list. Edit quantities to compute price and area.")

    up = st.file_uploader("Upload price list (xlsx/csv)", type=["xlsx", "csv"], key="pl_up")
    if up is not None:
        try:
            if up.name.lower().endswith(".xlsx"):
                df = pd.read_excel(up)
            else:
                df = pd.read_csv(up)
            cols_map = {c.lower(): c for c in df.columns}
            required = {"code", "name", "unit_price_eur"}
            if not required.issubset(set(cols_map.keys())):
                st.error("File must contain: code, name, unit_price_eur (optional area_m2).")
            else:
                order = ["code", "name", "unit_price_eur"] + (["area_m2"] if "area_m2" in cols_map else [])
                df = df[[cols_map[c] for c in order]]
                if "qty" not in df.columns:
                    df["qty"] = 0
                st.session_state.price_df = df.copy()
        except Exception as e:
            st.error(f"Failed to read file: {e}")

    edited = st.data_editor(
        st.session_state.price_df,
        use_container_width=True,
        num_rows="dynamic",
        column_config={"qty": st.column_config.NumberColumn("qty", min_value=0, step=1)},
        key="price_table"
    )

    if not edited.empty:
        unit = edited["unit_price_eur"].fillna(0)
        qty  = edited["qty"].fillna(0)
        total_price = float((unit * qty).sum())
        total_qty   = int(qty.astype(int).sum())
    else:
        total_price, total_qty = 0.0, 0
    st.session_state.taiga_total_qty = total_qty

    st.metric("Computed Taiga list price (€)", f"{total_price:,.0f}")

    total_area = None
    if not edited.empty and "area_m2" in edited.columns:
        try:
            total_area = float((edited["area_m2"].fillna(0) * edited["qty"].fillna(0)).sum())
        except Exception:
            total_area = 0.0

    qcol1, qcol2 = st.columns(2)
    with qcol1:
        st.metric("Selected Taiga quantity (units)", f"{total_qty}")
    with qcol2:
        if total_area is not None:
            st.metric("Area from products (m²)", f"{total_area:,.2f}")

    c1, c2 = st.columns(2)
    with c1:
        st.checkbox("Compute Taiga list price from selected products", key="override_price")
    with c2:
        if total_area is not None:
            st.checkbox("Override shared Area (m²) with total from products", key="override_area")

    if st.session_state.override_price:
        st.session_state.taiga_list_price = total_price

    st.session_state.area_from_products = total_area if total_area is not None else None

    if st.session_state.get("override_area") and st.session_state.get("area_from_products") is not None:
        st.info(f"Using area from products: {st.session_state.area_from_products:.2f} m²", icon="ℹ️")

def to_cycle_list(df: pd.DataFrame):
    """Convert the editable DataFrame to list[CyclePoint]. Accept % or decimals."""
    pts = []
    for r in df.itertuples(index=False):
        year = int(getattr(r, "year"))
        raw = getattr(r, "value_pct")
        try:
            v = float(raw)
            v = v / 100.0 if abs(v) > 1.0 else v
        except Exception:
            v = 0.0
        pts.append(CyclePoint(year=year, value_pct=v))
    return pts

def build_inputs_taiga(shared):
    """Build TCOInputs for Taiga using per-unit fields scaled by qty."""
    ss = st.session_state
    qty = int(ss.get("taiga_total_qty", 0))

    commissioning_total = float(ss.taiga_commissioning_cost_unit) * qty
    maint_total = float(ss.taiga_maint_total_unit) * qty
    dt_install_total_h = float(ss.taiga_dt_install_h_unit) * qty
    dt_maint_total_h   = float(ss.taiga_dt_maint_h_total_unit) * qty

    return TCOInputs(
        is_taiga=True,
        years=int(shared["years"]),
        wacc=float(shared["wacc"]),
        list_price=float(ss.taiga_list_price),
        area_m2=float(shared["area_m2"]),
        kwh_m2yr=float(shared["kwh_m2yr"]),
        elec_price=float(shared["elec_price"]),
        run_frac_trad=0.0,
        occ_rate=float(ss.taiga_occ_rate),
        standby_taiga=float(ss.taiga_standby),
        commissioning_cost=commissioning_total,
        commissioning_year=int(ss.taiga_commissioning_year),
        maint_total=maint_total,
        downtime_rate_per_hour=float(ss.taiga_dt_rate),
        downtime_hours_install=dt_install_total_h,
        downtime_hours_maint_total=dt_maint_total_h,
        eol_cost=float(ss.taiga_eol_cost),
        cycle_table=to_cycle_list(st.session_state.cycle_df),
        cycle_year=int(shared["cycle_year"]),
    )

def build_inputs_trad(shared):
    """Build TCOInputs for TRAD. Costs scaled by rooms (qty) and years where specified."""
    ss = st.session_state
    qty_rooms = int(ss.get("taiga_total_qty", 0))  # rooms = Taiga qty (can be separated later)

    # Investment = €/m² × area
    list_price_total = float(ss.trad_price_per_m2) * float(shared["area_m2"])

    # Commissioning & maintenance (scaled by rooms)
    commissioning_total = float(ss.trad_commissioning_cost_unit) * qty_rooms
    maint_total = float(ss.trad_maint_total_unit) * qty_rooms

    # Downtime hours (scaled by rooms); rate €/h unchanged
    dt_install_total_h = float(ss.trad_dt_install_h_unit) * qty_rooms
    dt_maint_total_h   = float(ss.trad_dt_maint_h_total_unit) * qty_rooms

    # End-of-life: % of investment
    eol_cost = float(ss.trad_eol_pct) * list_price_total

    return TCOInputs(
        is_taiga=False,
        years=int(shared["years"]),
        wacc=float(shared["wacc"]),
        list_price=list_price_total,
        area_m2=float(shared["area_m2"]),
        kwh_m2yr=float(shared["kwh_m2yr"]),
        elec_price=float(shared["elec_price"]),
        run_frac_trad=float(ss.get("trad_run_frac", 0.90)),
        occ_rate=0.0,
        standby_taiga=0.0,
        commissioning_cost=commissioning_total,
        commissioning_year=int(ss.trad_commissioning_year),
        maint_total=maint_total,
        downtime_rate_per_hour=float(ss.trad_dt_rate),
        downtime_hours_install=dt_install_total_h,
        downtime_hours_maint_total=dt_maint_total_h,
        eol_cost=eol_cost,
        cycle_table=[],
        cycle_year=0,
    )

def excel_bytes_all(taiga_summary, trad_summary, df_taiga, df_trad, df_delta, taiga_inp, trad_inp):
    """Build an Excel workbook with all outputs."""
    from openpyxl import Workbook
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary-Taiga"
    ws1.append(["Component", "PV (€)"])
    for k, v in taiga_summary.items():
        ws1.append([k, float(v)])

    ws2 = wb.create_sheet("Summary-TRAD")
    ws2.append(["Component", "PV (€)"])
    for k, v in trad_summary.items():
        ws2.append([k, float(v)])

    ws3 = wb.create_sheet("Summary-Delta (T-TR)")
    all_keys = sorted(set(taiga_summary.keys()) | set(trad_summary.keys()))
    ws3.append(["Component", "Delta PV (€)"])
    for k in all_keys:
        ws3.append([k, float(taiga_summary.get(k, 0.0) - trad_summary.get(k, 0.0))])

    for title, df in [("Yearly-Taiga", df_taiga), ("Yearly-TRAD", df_trad), ("Yearly-Delta", df_delta)]:
        ws = wb.create_sheet(title)
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append(list(r.values))

    ws4 = wb.create_sheet("CycleTable (Taiga only)")
    ws4.append(["Year", "Value%"])
    for cp in taiga_inp.cycle_table:
        ws4.append([cp.year, cp.value_pct])

    for title, inp in [("Inputs-Taiga", taiga_inp), ("Inputs-TRAD", trad_inp)]:
        ws = wb.create_sheet(title)
        for k, v in inp.__dict__.items():
            if k == "cycle_table":
                continue
            ws.append([k, v])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def _total_pv(summary: dict) -> float:
    """Return overall Total PV from a summary dict (fallback)."""
    if "total_pv" in summary:
        try:
            return float(summary["total_pv"])
        except Exception:
            pass
    return float(sum(v for v in summary.values() if isinstance(v, (int, float))))

# ---- Cost columns order & helpers for pivot ----
COST_COLS_ORDER = [
    "acquisition_pv",
    "buyback_pv",
    "commissioning_pv",
    "energy_cost_pv",
    "maintenance_pv",
    "downtime_pv",
    "eol_pv",
]

def _cycle_pct_for_year(year: int) -> float:
    """Read cycle % from st.session_state.cycle_df (accepts 50 or 0.50)."""
    try:
        df = st.session_state.cycle_df
        row = df.loc[df["year"] == year]
        if row.empty:
            return 0.0
        v = float(row["value_pct"].iloc[0])
        return v / 100.0 if abs(v) > 1.0 else v
    except Exception:
        return 0.0

def ensure_cost_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure all known cost columns exist, keep fixed column order."""
    if df is None or df.empty:
        return df
    for col in COST_COLS_ORDER:
        if col not in df.columns:
            df[col] = 0.0
    other = [c for c in df.columns if c not in (["year"] + COST_COLS_ORDER)]
    return df[["year"] + COST_COLS_ORDER + other]

def ensure_year_row(df: pd.DataFrame, year: int) -> pd.DataFrame:
    """Ensure there is a row for given year; if missing, append a zero row with all cost columns."""
    if df is None or df.empty:
        base = pd.DataFrame({"year": [year]})
        for c in COST_COLS_ORDER:
            base[c] = 0.0
        return base
    if "year" not in df.columns:
        df = df.copy()
        df["year"] = 0
    if (df["year"] == year).any():
        return df
    zero_row = {col: 0.0 for col in COST_COLS_ORDER if col in df.columns}
    zero_row["year"] = year
    df = pd.concat([df, pd.DataFrame([zero_row])], ignore_index=True)
    return df.sort_values("year").reset_index(drop=True)

def pivot_for_display(df, value_cols=None):
    """Rows = cost items, Cols = Year 0..N + Total; includes per-year total row."""
    if df is None or df.empty:
        return pd.DataFrame()
    default_cols = COST_COLS_ORDER.copy()
    if value_cols is None:
        value_cols = default_cols
    value_cols = [c for c in value_cols if c in df.columns]
    if not value_cols:
        return pd.DataFrame()
    pivoted = df.set_index("year")[value_cols].T
    pivoted.index.name = "Cost item"
    pivoted.columns = [f"Year {c}" for c in pivoted.columns]
    # Friendly names
    name_map = {
        "acquisition_pv": "Acquisition",
        "buyback_pv": "Buyback (−)",
        "commissioning_pv": "Commissioning",
        "energy_cost_pv": "Operation (Energy)",
        "maintenance_pv": "Maintenance",
        "downtime_pv": "Downtime",
        "eol_pv": "End-of-Life",
    }
    pivoted.rename(index=name_map, inplace=True)
    # Row total (per item)
    pivoted["Total"] = pivoted.sum(axis=1)
    # Per-year total row
    year_cols = [c for c in pivoted.columns if c.startswith("Year ")]
    per_year_total = pivoted[year_cols].sum(axis=0)
    per_year_total["Total"] = per_year_total.sum()
    pivoted.loc["— Total (per year) —"] = per_year_total
    return pivoted

def pv_total_from_yearly(df: pd.DataFrame) -> float:
    """Sum present value over all cost columns and years to get Overall Total PV."""
    if df is None or df.empty:
        return 0.0
    cols = [c for c in COST_COLS_ORDER if c in df.columns]
    return float(df[cols].sum().sum())

def component_summary_from_yearly(df: pd.DataFrame) -> dict:
    """Return per-component PV and total_pv from yearly table."""
    if df is None or df.empty:
        return {"total_pv": 0.0}
    cols = [c for c in COST_COLS_ORDER if c in df.columns]
    comp = df[cols].sum().to_dict()
    comp["total_pv"] = float(sum(comp.values()))
    return comp

# -------------------- UI --------------------
img_bytes = Path("logo.PNG").read_bytes()
b64 = base64.b64encode(img_bytes).decode()
css = css.replace("URL_LOGO_PLACEHOLDER", f"data:image/png;base64,{b64}")
st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)

init_state()
st.title("Taiga Concept Owner Calculator")

with st.expander("Shared parameters", expanded=False):
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.number_input("Horizon (years)", 1, 50, key="years")
    with c2:
        st.number_input("WACC (decimal, e.g., 0.08)", 0.0, 1.0, key="wacc", step=0.01)
    with c3:
        st.number_input("Area (m²)", 0.0, 1e9, key="area_m2", step=1.0)
    with c4:
        st.number_input("kWh / m² / year", 0.0, 1e9, key="kwh_m2yr", step=1.0)
    c5, c6 = st.columns(2)
    with c5:
        st.number_input("Electricity price (€/kWh)", 0.0, 10.0, key="elec_price", step=0.01)
    with c6:
        st.caption("Cycle settings are under Taiga section.")

tab_products, tab_leasing = st.tabs(["Products", "Leasing"])
with tab_products:
    st.markdown("### Products and Parameters")
    colT, colR = st.columns(2, gap="large")

    with colT:
        st.subheader("Taiga")

        with st.expander("Taiga Pricing", expanded=False):
            taiga_price_list_ui()

        with st.expander("Taiga Parameters", expanded=False):
            st.number_input("List price (€)", 0.0, 1e12, key="taiga_list_price",
                            step=1000.0, disabled=st.session_state.get("override_price", False))
            st.slider("Occupancy rate", 0.0, 1.0, key="taiga_occ_rate")
            st.slider("Standby share", 0.0, 1.0, key="taiga_standby")
            st.number_input("Commissioning cost per unit (€)", 0.0, 1e12,
                            key="taiga_commissioning_cost_unit", step=50.0)
            st.number_input("Maintenance per unit annually (€)", 0.0, 1e12,
                            key="taiga_maint_total_unit", step=10.0)
            c1, c2, c3 = st.columns(3)
            with c1:
                st.number_input("Downtime rate (€/h)", 0.0, 1e9, key="taiga_dt_rate", step=1.0)
            with c2:
                st.number_input("Install downtime per unit (h)", 0.0, 1e9,
                                key="taiga_dt_install_h_unit", step=0.5)
            with c3:
                st.number_input("Maint downtime per unit annually (h)", 0.0, 1e9,
                                key="taiga_dt_maint_h_total_unit", step=0.5)
            st.number_input("Commissioning year", 0, 50, key="taiga_commissioning_year")
            st.number_input("End-of-life cost (total) (€)", 0.0, 1e12, key="taiga_eol_cost", step=100.0)

        qty = st.session_state.get("taiga_total_qty", 0)
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            st.metric("Commissioning total (€)",
                      f"{st.session_state.taiga_commissioning_cost_unit * qty:,.0f}")
        with d2:
            st.metric("Maintenance total (€)",
                      f"{st.session_state.taiga_maint_total_unit * qty * int(st.session_state.years):,.0f}")
        with d3:
            st.metric("Install downtime total (h)",
                      f"{st.session_state.taiga_dt_install_h_unit * qty:,.1f}")
        with d4:
            st.metric("Maint downtime total (h)",
                      f"{st.session_state.taiga_dt_maint_h_total_unit * qty * int(st.session_state.years):,.0f}")

        with st.expander("Taiga Buyback", expanded=False):
            st.markdown("### Taiga buyback (Cycle)")
            st.number_input("Cycle (buyback) year (Taiga)", 0, 50, key="cycle_year")
            st.markdown("**Taiga Cycle table (Year, Value %)**")
            st.caption("Edit or add rows; values can be 70 (%=0.70) or 0.70.")
            st.session_state.cycle_df = st.data_editor(
                st.session_state.cycle_df,
                num_rows="dynamic",
                use_container_width=True
            )

    effective_area = (
        st.session_state.area_from_products
        if (st.session_state.get("override_area") and st.session_state.get("area_from_products") is not None)
        else st.session_state.area_m2
    )

    with colR:
        st.subheader("Traditional Building (TRAD)")
        with st.expander("Building Pricing", expanded=False):
            st.number_input("Investment price per m² (TRAD) (€)", 0.0, 1e6,
                            key="trad_price_per_m2", step=50.0)
        with st.expander("Building Parameters", expanded=False):
            st.slider("Run fraction (Traditional)", 0.0, 1.0, key="trad_run_frac")
            eol_pct_input = st.number_input("End-of-life (% of investment)", 0.0, 100.0,
                                            value=st.session_state.trad_eol_pct * 100, step=1.0)
            st.session_state.trad_eol_pct = float(eol_pct_input) / 100.0
            st.number_input("Commissioning cost per room (€)", 0.0, 1e9,
                            key="trad_commissioning_cost_unit", step=10.0)
            st.number_input("Maintenance total per room annually (€)", 0.0, 1e12,
                            key="trad_maint_total_unit", step=100.0)
            c1, c2, c3 = st.columns(3)
            with c1:
                st.number_input("Downtime rate (€/h)", 0.0, 1e9, key="trad_dt_rate", step=1.0)
            with c2:
                st.number_input("Install downtime per room (h)", 0.0, 1e9,
                                key="trad_dt_install_h_unit", step=1.0)
            with c3:
                st.number_input("Maint downtime per room annually (h)", 0.0, 1e9,
                                key="trad_dt_maint_h_total_unit", step=1.0)
            st.number_input("Commissioning year (kept for API)", 0, 50, key="trad_commissioning_year")

        qty_rooms = st.session_state.get("taiga_total_qty", 0)
        trad_invest_total = float(st.session_state.trad_price_per_m2) * float(effective_area)
        trad_comm_total   = float(st.session_state.trad_commissioning_cost_unit) * qty_rooms
        trad_maint_total  = float(st.session_state.trad_maint_total_unit) * qty_rooms * int(st.session_state.years)
        trad_dt_install_h_total = float(st.session_state.trad_dt_install_h_unit) * qty_rooms
        trad_dt_maint_h_total   = float(st.session_state.trad_dt_maint_h_total_unit) * qty_rooms * int(st.session_state.years)
        trad_eol_total = float(st.session_state.trad_eol_pct) * trad_invest_total

        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("Computed TRAD investment (€)", f"{trad_invest_total:,.0f}")
        with m2:
            st.metric("Rooms (qty) used", f"{qty_rooms}")
        with m3:
            st.metric("Years", f"{int(st.session_state.years)}")

        m4, m5, m6, m7 = st.columns(4)
        with m4:
            st.metric("Commissioning total (€)", f"{trad_comm_total:,.0f}")
        with m5:
            st.metric("Maintenance total (€)", f"{trad_maint_total:,.0f}")
        with m6:
            st.metric("Install downtime total (h)", f"{trad_dt_install_h_total:,.1f}")
        with m7:
            st.metric("Maint downtime total (h)", f"{trad_dt_maint_h_total:,.1f}")
        st.metric("End-of-life total (€)", f"{trad_eol_total:,.0f}")

with tab_leasing:
    st.subheader("Leasing")

    with st.expander("Leasing Costs", expanded=False):
        st.markdown("#### Monthly factor table")
        # Default table (editable in UI)

        if "lease_factors_df" not in st.session_state:
            st.session_state.lease_factors_df = pd.DataFrame({
                "term_years": [3, 4, 5, 6, 7],
                "monthly_factor": [1.95, 1.70, 1.55, 1.45, 1.38],  # %/month, user may also enter as decimal
            })
    
        lf_edit = st.data_editor(
            st.session_state.lease_factors_df,
            use_container_width=True,
            num_rows="dynamic",
            key="lease_factors_editor",
            column_config={
                "term_years": st.column_config.NumberColumn("term_years", min_value=1, step=1),
                "monthly_factor": st.column_config.NumberColumn(
                    "monthly_factor",
                    help="Enter as % per month (e.g., 1.55) or decimal (0.0155)."
                ),
            }
        )
        st.session_state.lease_factors_df = lf_edit.copy()

    
    with st.expander("Change Default Inputs", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            lease_term = st.number_input("Leasing term (years)", 1, 15, value=int(st.session_state.years), step=1)
        with c2:
            # Use the same WACC definition (annual)
            wacc_lease = st.number_input("WACC (annual, decimal)", 0.0, 1.0, value=float(st.session_state.wacc), step=0.01)
        with c3:
            taiga_price_default = float(st.session_state.taiga_list_price)
            taiga_price = st.number_input("Base price (€)", 0.0, 1e12, value=taiga_price_default, step=100.0)
        with c4:
            cycle_year_input = st.number_input("Buyback year (0 = none)", 0, 50, value=int(st.session_state.cycle_year), step=1)

    # Convert Streamlit cycle_df -> leasing_calc.CyclePoint list
    def _to_cp_list(df):
        out = []
        if df is not None and not df.empty:
            for r in df.itertuples(index=False):
                year = int(getattr(r, "year"))
                raw = getattr(r, "value_pct")
                try:
                    v = float(raw)
                    v = v/100.0 if abs(v) > 1.0 else v
                except Exception:
                    v = 0.0
                out.append(lc.CyclePoint(year=year, value_pct=v))
        return out

    cp_list = _to_cp_list(st.session_state.cycle_df)

    # Get monthly factor for the chosen term
    mo_factor_raw = lc.monthly_factor_for_term(lease_term, st.session_state.lease_factors_df)
    # Be robust: accept factor both as %/month (e.g., 1.55) and as decimal (0.0155)
    mo_factor = mo_factor_raw / 100.0 if mo_factor_raw > 1 else mo_factor_raw

    # Monthly payments (base + with buyback)
    base_mo, mo_with_buyback = lc.monthly_payment_with_buyback(
        list_price=taiga_price,
        monthly_factor=mo_factor,
        wacc_annual=wacc_lease,     # same WACC definition
        term_years=lease_term,
        cycle_year=cycle_year_input,
        cycle_table=cp_list
    )

    st.markdown("### Overall Taiga Leasing Costs")

    cA, cB, cC = st.columns(3)

    with cA:
        st.metric("Monthly factor", f"{mo_factor * 100:.2f}")

    with cB:
        st.metric("Base monthly (€)", f"{base_mo:,.0f}")

    with cC:
        st.metric(
            "Monthly, with buyback (€)",
            f"{mo_with_buyback:,.0f}",
            delta=f"{mo_with_buyback - base_mo:,.0f} € vs base",
            delta_color="inverse"  # vihreä jos buyback pienempi kuin base
        )

    with st.expander("Detailed Leasing Results", expanded=False):
            st.markdown("#### Yearly PV (leasing stream)")
            df_leasing_yearly = lc.leasing_yearly_pv_table(
                list_price=taiga_price,
                monthly_factor=mo_factor,
                wacc_annual=wacc_lease,     # same WACC definition
                term_years=lease_term,
                cycle_year=cycle_year_input,
                cycle_table=cp_list
            )
            st.dataframe(df_leasing_yearly.round(0), use_container_width=True)

            st.markdown("#### Leasing pivot")
            pv_leasing = lc.pivot_leasing_for_display(df_leasing_yearly).round(0)
            st.dataframe(pv_leasing, use_container_width=True)

            st.markdown("#### Summary")
            st.write(f"- **Monthly payment (leasing)**: {base_mo:,.0f} €")
            st.write(f"- **Monthly payment (with buyback)**: {mo_with_buyback:,.0f} €")

# Area info line
st.caption(
        f"Using area for calculations: "
        f"{(st.session_state.area_from_products if st.session_state.get('override_area') and st.session_state.get('area_from_products') is not None else st.session_state.area_m2):.2f} m²"
)

# -------------------- Build inputs & compute yearly tables --------------------

shared = dict(
    years=st.session_state.years,
    wacc=st.session_state.wacc,
    area_m2=effective_area,
    kwh_m2yr=st.session_state.kwh_m2yr,
    elec_price=st.session_state.elec_price,
    cycle_year=st.session_state.cycle_year,
)
taiga_inp = build_inputs_taiga(shared)
trad_inp  = build_inputs_trad(shared)

# (You may still compute raw summaries if needed elsewhere)
# taiga_sum_raw = dict(compute_tco(taiga_inp))
# trad_sum_raw  = dict(compute_tco(trad_inp))

# Yearly rows from engine
taiga_rows = yearly_breakdown(taiga_inp)
trad_rows  = yearly_breakdown(trad_inp)

df_taiga = pd.DataFrame(taiga_rows)
df_trad  = pd.DataFrame(trad_rows)

# ---- Inject Acquisition & Buyback into yearly DataFrames ----

# Ensure base cost columns exist first
df_taiga = ensure_cost_columns(df_taiga)
df_trad  = ensure_cost_columns(df_trad)

# Ensure Year 0 exists, then set acquisition at Year 0 (no discount)
df_taiga = ensure_year_row(df_taiga, 0)
df_trad  = ensure_year_row(df_trad, 0)

if "acquisition_pv" not in df_taiga.columns:
    df_taiga["acquisition_pv"] = 0.0
if "acquisition_pv" not in df_trad.columns:
    df_trad["acquisition_pv"] = 0.0

df_taiga.loc[df_taiga["year"] == 0, "acquisition_pv"] = float(taiga_inp.list_price)
df_trad.loc[df_trad["year"] == 0, "acquisition_pv"]   = float(trad_inp.list_price)

# Taiga Buyback at cycle_year (discounted by WACC) — ensure the row exists first
yb = int(getattr(taiga_inp, "cycle_year", 0))
if yb > 0:
    pct = _cycle_pct_for_year(yb)  # e.g., 0.30
    if pct > 0:
        df_taiga = ensure_year_row(df_taiga, yb)
        if "buyback_pv" not in df_taiga.columns:
            df_taiga["buyback_pv"] = 0.0
        discount = (1.0 + float(taiga_inp.wacc)) ** yb
        df_taiga.loc[df_taiga["year"] == yb, "buyback_pv"] = - float(taiga_inp.list_price) * pct / discount

# Re-ensure fixed order & zero-fill (in case we added columns)
df_taiga = ensure_cost_columns(df_taiga)
df_trad  = ensure_cost_columns(df_trad)

# ---- Delta (include acquisition & buyback) ----
df_delta = None
if not df_taiga.empty and not df_trad.empty:
    common_cols = [c for c in df_taiga.columns if c in df_trad.columns]
    sum_cols = [c for c in common_cols if c != "year"]
    merged = df_taiga[["year"] + sum_cols].merge(
        df_trad[["year"] + sum_cols], on="year", suffixes=("_T", "_TR")
    )
    for c in sum_cols:
        merged[c] = merged[f"{c}_T"] - merged[f"{c}_TR"]
    df_delta = merged[["year"] + sum_cols]

# ---- Summaries derived from yearly tables (single source of truth) ----
taiga_sum = component_summary_from_yearly(df_taiga)
trad_sum  = component_summary_from_yearly(df_trad)

# Overall Total PV metrics (identical to pivot totals)
taiga_total_pv = pv_total_from_yearly(df_taiga)
trad_total_pv  = pv_total_from_yearly(df_trad)
delta_total_pv = taiga_total_pv - trad_total_pv

# -------------------- Results & Pivots --------------------

with st.expander("Total Ownership Costs (no financing costs)", expanded=False):
    st.markdown("---")
    st.subheader("Summary (Present Value €)")

    cols = st.columns(3)
    with cols[0]:
        st.caption("Taiga")
        st.dataframe(pd.DataFrame([taiga_sum]), use_container_width=True, hide_index=True)
    with cols[1]:
        st.caption("TRAD")
        st.dataframe(pd.DataFrame([trad_sum]), use_container_width=True, hide_index=True)
    with cols[2]:
        st.caption("Delta (Taiga − TRAD)")
        keys = sorted(set(taiga_sum.keys()) | set(trad_sum.keys()))
        delta = {k: taiga_sum.get(k, 0.0) - trad_sum.get(k, 0.0) for k in keys}
        st.dataframe(pd.DataFrame([delta]), use_container_width=True, hide_index=True)

    st.markdown("### Yearly breakdowns")

    colA, colB, colC = st.columns(3)

    def _show_pivot(df, title):
        """Render the cost pivot (rows=cost items, cols=years)."""
        st.caption(title)
        pv = pivot_for_display(df).round(0)
        if pv.empty:
            st.info("No data.", icon="ℹ️")
            return
        st.dataframe(
            pv,
            use_container_width=True,
            column_config={
                col: st.column_config.NumberColumn(col, format="%.0f", width="small")
                for col in pv.columns
            }
        )

    with colA:
        _show_pivot(df_taiga, "Taiga")
    with colB:
        _show_pivot(df_trad, "TRAD")
    with colC:
        if df_delta is not None and not df_delta.empty:
            _show_pivot(df_delta, "Delta (Taiga − TRAD)")
        else:
            st.caption("Delta (Taiga − TRAD)")
            st.info("Delta becomes available when both sides have rows.", icon="ℹ️")

# -------------------- Overall Total PV metrics --------------------

st.markdown("### Overall Total Cost Comparison")
m1, m2, m3 = st.columns(3)
with m1:
    st.metric("Taiga Total PV (€)", f"{taiga_total_pv:,.0f}")
with m2:
    st.metric("TRAD Total PV (€)", f"{trad_total_pv:,.0f}")
with m3:
    st.metric("Delta Total PV (€)  (Taiga − TRAD)", f"{delta_total_pv:,.0f}",
              delta=f"{delta_total_pv:,.0f}")

# -------------------- Export --------------------


# Build pivots for Word export
pv_taiga = pivot_for_display(df_taiga).round(0)
pv_trad  = pivot_for_display(df_trad).round(0)
pv_delta = pivot_for_display(df_delta).round(0) if df_delta is not None and not df_delta.empty else None

payload = {
    "customer_name": "Demo Customer",
    "project_name": "Demo Project",
    "date_str": pd.Timestamp.now().strftime("%Y-%m-%d"),
    "params": shared,
    "results": {
        "TCO_TRAD_PV": trad_total_pv,
        "TCO_TAIGA_PV": taiga_total_pv,
        "DIFF_TRAD_TAIGA": trad_total_pv - taiga_total_pv,
    },
}

# ensure latest proposal_doc in memory
reload(proposal_doc)

doc_bytes = proposal_doc.generate_proposal_doc(
    payload=payload,
    df_pivot_taiga=pv_taiga,
    df_pivot_trad=pv_trad,
    df_pivot_delta=pv_delta,
    locale="fi_FI",
    logo_path="logo.png"  # e.g., "logo.png"
)

st.download_button(
    label="📄 Download Word Summary",
    data=doc_bytes,
    file_name="TCO_Summary.docx",
    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
)

# -------------------- End --------------------